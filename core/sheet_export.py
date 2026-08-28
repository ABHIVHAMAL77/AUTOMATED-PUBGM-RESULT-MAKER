"""Automatic tournament sheet export — replaces the manual JSON→CSV workflow.

After every finalized match this module rebuilds, from the saved match JSONs:

    data/exports/Tournament Sheet.xlsx   (tabs mirror the Google Sheet layout)
        SETUP             tournament name, stage, matches, teams, point table
        OVERALL GD        raw per-player API rows of every match
        Match Standings   per-match results, one block per match
        Overall Standings cumulative standings after the latest match
        Player Stats      aggregated player leaderboard with MVP rating

    data/exports/csv/    the same data as plain CSV files (easy to import
                         into Google Sheets or use anywhere else)

Everything is rebuilt from data/results/match_*.json each time, so the sheet
is always consistent even if a match is re-saved or deleted.
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Team slots shown in the SETUP tab (slot number = API team ID)
MAX_SLOTS = 25

# Raw API columns for the OVERALL GD tab, in the same order as the user's
# original Google Sheet.
GD_API_COLUMNS = [
    "uId", "playerName", "playerOpenId", "picUrl", "showPicUrl", "teamId",
    "teamName", "character", "isFiring", "bHasDied", "location__x",
    "location__y", "location__z", "health", "healthMax", "liveState",
    "killNum", "killNumBeforeDie", "playerKey", "gotAirDropNum",
    "maxKillDistance", "damage", "killNumInVehicle", "killNumByGrenade",
    "AIKillNum", "BossKillNum", "rank", "isOutsideBlueCircle", "inDamage",
    "heal", "headShotNum", "survivalTime", "driveDistance", "marchDistance",
    "assists", "outsideBlueCircleTime", "knockouts", "rescueTimes",
    "useSmokeGrenadeNum", "useFragGrenadeNum", "useBurnGrenadeNum",
    "useFlashGrenadeNum", "PoisonTotalDamage", "UseSelfRescueTime",
    "UseEmergencyCallTime",
]

HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(bold=True, color="F0C24B", name="Arial")
TITLE_FONT = Font(bold=True, size=14, name="Arial")
BODY_FONT = Font(name="Arial")


def _location_fields(raw: dict) -> dict:
    """The API nests location as {"location": {"x": ..}}; flatten it."""
    loc = raw.get("location") or {}
    return {
        "location__x": raw.get("location__x", loc.get("x", "")),
        "location__y": raw.get("location__y", loc.get("y", "")),
        "location__z": raw.get("location__z", loc.get("z", "")),
    }


def _fmt_minutes(seconds) -> str:
    try:
        s = int(float(seconds))
    except (TypeError, ValueError):
        return ""
    return f"{s // 60}:{s % 60:02d}"


def _gd_rows(matches: list) -> list:
    """Flatten all matches into OVERALL GD rows (header + data)."""
    header = ["Match", "Team ID", "Team Placement"] + GD_API_COLUMNS + ["Survival Time (In minutes)"]
    rows = [header]
    for match in matches:
        for r in match.get("results", []):
            for p in r.get("players", []):
                raw = dict(p.get("raw") or {})
                raw.update(_location_fields(raw))
                # fall back to the flat stats saved by scoring when the raw
                # API row is unavailable (e.g. mock data)
                fallback = {
                    "uId": p.get("uId"), "playerName": p.get("playerName"),
                    "teamId": r.get("teamId"), "teamName": r.get("teamName"),
                    "killNum": p.get("kills"), "damage": p.get("damage"),
                    "knockouts": p.get("knockouts"), "headShotNum": p.get("headshots"),
                    "assists": p.get("assists"), "inDamage": p.get("damageReceived"),
                    "survivalTime": p.get("survivalTime"), "heal": p.get("heal"),
                    "rescueTimes": p.get("rescues"), "maxKillDistance": p.get("longestKill"),
                    "killNumByGrenade": p.get("grenadeKills"), "rank": r.get("placement"),
                }
                row = [match.get("matchNumber"), r.get("teamId"), r.get("placement")]
                for col in GD_API_COLUMNS:
                    val = raw.get(col, fallback.get(col, ""))
                    if isinstance(val, dict):
                        val = str(val)
                    row.append(val)
                row.append(_fmt_minutes(raw.get("survivalTime", p.get("survivalTime"))))
                rows.append(row)
    return rows


def _match_standing_rows(matches: list) -> list:
    rows = [["Match", "Map", "Rank", "Team ID", "Team Name", "Place Points",
             "Elims", "Elim Points", "Total Points", "WWCD"]]
    for match in matches:
        for r in match.get("results", []):
            rows.append([
                match.get("matchNumber"), match.get("map", ""), r["placement"],
                r["teamId"], r["teamName"], r["placementPoints"], r["kills"],
                r["killPoints"], r["totalPoints"], 1 if r.get("wwcd") else 0,
            ])
    return rows


def _overall_rows(standings: list) -> list:
    rows = [["Rank", "Team Name", "WWCD", "Eliminations", "Placement Points",
             "Elim Points", "Total Points", "Matches Played"]]
    for a in standings:
        rows.append([a["rank"], a["teamName"], a["wwcd"], a["kills"],
                     a["placementPoints"], a["killPoints"], a["totalPoints"],
                     a["matches"]])
    return rows


def _player_rows(stats: list) -> list:
    rows = [["Rank", "Player UID", "Player IGN", "Team Name", "Matches",
             "Elims", "Damage", "Headshots", "Assists", "Knockouts",
             "Dmg Received", "Survival (sec)", "Heal", "Rescues",
             "Longest Elimination", "Grenade Elims", "Contribution", "MVP Rating"]]
    for a in stats:
        rows.append([a["rank"], a["uId"], a["playerName"], a["teamName"],
                     a["matches"], a["kills"], a["damage"], a["headshots"],
                     a["assists"], a["knockouts"], a["damageReceived"],
                     a["survivalTime"], a["heal"], a["rescues"],
                     a["longestKill"], a["grenadeKills"], a["contribution"],
                     a["mvpRating"]])
    return rows


def _setup_rows(event: dict) -> list:
    ps = event.get("pointSystem", {})
    rows = [
        ["TOURNAMENT NAME", event.get("eventName", "")],
        ["STAGE", event.get("stage", "")],
        ["TOTAL MATCHES", event.get("totalMatches", "")],
        [],
        ["Place", "Points"],
    ]
    for i, pts in enumerate(ps.get("placementPoints", []), start=1):
        rows.append([i, pts])
    rows.append([f"{len(ps.get('placementPoints', []))+1}+", 0])
    rows.append(["Per Elimination", ps.get("killPoint", 1)])
    rows.append([])
    rows.append(["Slot No", "Team Name", "Team Tag"])
    by_slot = {int(t.get("teamId", 0)): t for t in event.get("teams", [])}
    for slot in range(1, MAX_SLOTS + 1):
        t = by_slot.get(slot, {})
        rows.append([slot, t.get("teamName", ""), t.get("shortName", "")])
    return rows


def _write_sheet(ws, rows, header_row_indexes=(0,)):
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri + 1, column=ci, value=val)
            cell.font = BODY_FONT
            if ri in header_row_indexes:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
    # readable column widths
    for ci in range(1, (max(len(r) for r in rows) if rows else 0) + 1):
        best = max((len(str(r[ci - 1])) for r in rows if len(r) >= ci), default=8)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(best + 2, 8), 34)
    if rows and len(rows) > 1:
        ws.freeze_panes = "A2"


def _write_csv(path: Path, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)


def export_tournament_sheet(event_manager) -> Path:
    """Rebuild the xlsx workbook + CSVs from all saved matches."""
    em = event_manager
    matches = [em.load_match(n) for n in em.list_match_numbers()]
    matches = [m for m in matches if m]
    standings = em.overall_standings()
    players = em.player_stats()

    gd = _gd_rows(matches)
    match_rows = _match_standing_rows(matches)
    overall = _overall_rows(standings)
    player_rows = _player_rows(players)
    setup = _setup_rows(em.event)

    wb = Workbook()
    ws = wb.active
    ws.title = "SETUP"
    _write_sheet(ws, setup, header_row_indexes=())
    ws["A1"].font = TITLE_FONT
    _write_sheet(wb.create_sheet("OVERALL GD"), gd)
    _write_sheet(wb.create_sheet("Match Standings"), match_rows)
    _write_sheet(wb.create_sheet("Overall Standings"), overall)
    _write_sheet(wb.create_sheet("Player Stats"), player_rows)

    xlsx_path = em.exports_dir / "Tournament Sheet.xlsx"
    wb.save(xlsx_path)

    csv_dir = em.exports_dir / "csv"
    csv_dir.mkdir(exist_ok=True)
    _write_csv(csv_dir / "overall_gd.csv", gd)
    _write_csv(csv_dir / "match_standings.csv", match_rows)
    _write_csv(csv_dir / "overall_standings.csv", overall)
    _write_csv(csv_dir / "player_stats.csv", player_rows)

    return xlsx_path
