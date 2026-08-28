"""Headless smoke test: simulate 2 mock matches end-to-end and render PNGs."""

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from core.mock_data import MockDataGenerator
from core.match_tracker import MatchTracker
from core.event_manager import EventManager
from core.scoring import DEFAULT_POINT_SYSTEM
from core import result_graphic
from core.sheet_export import export_tournament_sheet

tmp = Path(tempfile.mkdtemp(prefix="pubgm_test_"))
events = EventManager(tmp)
events.event["eventName"] = "Test Cup 2026"
events.event["stage"] = "Open Qualifiers"
events.event["totalMatches"] = 30

for match_num in (1, 2):
    gen = MockDataGenerator()
    tracker = MatchTracker()
    for tick in range(200):
        snap = gen.fetch()
        tracker.update(snap.team_states())
        if tracker.is_match_over:
            print(f"Match {match_num}: over after {tick + 1} ticks, "
                  f"alive teams = {tracker.alive_team_count}")
            break
    assert tracker.is_match_over, "match never ended"
    results = tracker.build_results(DEFAULT_POINT_SYSTEM)
    assert len(results) == 16, f"expected 16 teams, got {len(results)}"
    placements = [r["placement"] for r in results]
    assert placements == list(range(1, 17)), f"bad placements: {placements}"
    assert results[0]["wwcd"] is True
    assert results[0]["placementPoints"] == 10
    assert results[1]["placementPoints"] == 6
    assert results[8]["placementPoints"] == 0
    for r in results:
        assert r["totalPoints"] == r["placementPoints"] + r["kills"] * 1
    events.save_match_result(match_num, "Erangel", results)
    print(f"  winner: {results[0]['teamName']} "
          f"({results[0]['kills']} elims, {results[0]['totalPoints']} pts)")

standings = events.overall_standings()
assert len(standings) == 16
assert standings[0]["rank"] == 1
total_check = sum(a["matches"] for a in standings)
assert total_check == 32, total_check
print(f"Overall leader: {standings[0]['teamName']} with {standings[0]['totalPoints']} pts, "
      f"{standings[0]['wwcd']} WWCD")

m1 = events.load_match(1)
p1 = result_graphic.render_match_results(m1, events.exports_dir / "match_01_results.png")
p2 = result_graphic.render_overall_standings(
    standings, "Test Cup 2026", 2, events.exports_dir / "overall_standings.png")
print(f"PNG 1: {p1} ({p1.stat().st_size} bytes)")
print(f"PNG 2: {p2} ({p2.stat().st_size} bytes)")

# --- player stats -----------------------------------------------------------
stats = events.player_stats()
assert len(stats) == 64, f"expected 64 players, got {len(stats)}"
assert all(a["matches"] == 2 for a in stats)
assert stats[0]["kills"] >= stats[-1]["kills"], "not sorted by elims"
total_kills_players = sum(a["kills"] for a in stats)
total_kills_teams = sum(a["kills"] for a in standings)
assert total_kills_players == total_kills_teams, "player/team kill mismatch"
mvp_sum = sum(a["mvpRating"] for a in stats)
assert 0.99 < mvp_sum < 1.01, f"MVP ratings should sum to ~1, got {mvp_sum}"
top = stats[0]
print(f"Top player: {top['playerName']} ({top['teamName']}) — {top['kills']} elims, "
      f"{top['damage']} dmg, MVP {top['mvpRating']:.4f}")

# --- tournament sheet export -------------------------------------------------
xlsx = export_tournament_sheet(events)
wb = load_workbook(xlsx)
assert wb.sheetnames == ["SETUP", "OVERALL GD", "Match Standings",
                         "Overall Standings", "Player Stats"], wb.sheetnames
gd = wb["OVERALL GD"]
assert gd.max_row == 1 + 64 * 2, f"GD rows: {gd.max_row}"   # header + 64 players x 2 matches
ov = wb["Overall Standings"]
assert ov.max_row == 17
ps = wb["Player Stats"]
assert ps.max_row == 65
ms = wb["Match Standings"]
assert ms.max_row == 1 + 16 * 2
for name in ("overall_gd", "match_standings", "overall_standings", "player_stats"):
    csv_file = events.exports_dir / "csv" / f"{name}.csv"
    assert csv_file.exists() and csv_file.stat().st_size > 100, csv_file
print(f"Tournament sheet: {xlsx} ({xlsx.stat().st_size} bytes) — all 5 tabs verified")
print("CSV files verified")
print("ALL TESTS PASSED")


def test_pipeline_smoke_script_completed():
    pass
